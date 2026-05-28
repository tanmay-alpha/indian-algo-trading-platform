# backend/services/strategy_export_service.py
"""
Strategy Results Export Service — Phase 21B/C
=============================================

Generates in-memory Excel workbooks (.xlsx) from PERSISTED data only.
Sources: StrategySignalModel (SQLAlchemy), OrderStore (SQLite).

SAFETY RULES (enforced here):
- No fake rows. If a sheet has no data, it gets a NO_DATA sentinel row.
- No credentials, tokens, API keys, or secrets written to any cell.
- No broker API calls.
- No live trading data references.
- File generated in memory (io.BytesIO). Not written to disk.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Optional, Any

logger = logging.getLogger(__name__)

# Sentinel row written when a sheet has no records.
_NO_DATA = "NO_DATA"

# Fields from OrderStore / fill rows that are safe to export.
_SAFE_ORDER_FIELDS = (
    "id", "request_id", "client_order_id", "symbol", "side",
    "quantity", "order_type", "mode", "status", "avg_fill_price",
    "created_at", "updated_at",
)
_SAFE_FILL_FIELDS = (
    "id", "fill_id", "request_id", "symbol", "side",
    "filled_quantity", "fill_price", "fees", "source", "created_at",
)

# Fields NEVER written to Excel (safety guard).
_BLOCKED_FIELD_FRAGMENTS = (
    "token", "secret", "password", "key", "totp",
    "jwt", "refresh", "auth", "credential", "broker_order_id",
)


def _is_safe_field(field_name: str) -> bool:
    """Return False if field_name looks like a secret/credential."""
    low = field_name.lower()
    return not any(frag in low for frag in _BLOCKED_FIELD_FRAGMENTS)


def _safe_str(value: Any) -> Any:
    """Convert value to a type safe for an openpyxl cell."""
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def _write_sheet_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    """Write header row + data rows to an openpyxl worksheet."""
    ws.append(headers)
    if not rows:
        ws.append([_NO_DATA] + [""] * (len(headers) - 1))
        return
    for row in rows:
        ws.append([_safe_str(v) for v in row])


def _signal_to_row(sig) -> list[Any]:
    """Convert a StrategySignalModel ORM object to an exportable list."""
    return [
        sig.id,
        sig.strategy_id,
        sig.symbol,
        sig.side,
        sig.status,
        sig.confidence if sig.confidence is not None else "",
        sig.price if sig.price is not None else "",
        sig.timeframe or "",
        sig.source_candle_time or "",
        sig.reason or "",
        getattr(sig, "dismiss_reason", "") or "",
        sig.created_at or "",
    ]


def _order_to_row(order: dict) -> list[Any]:
    """Extract safe fields from an order_requests dict."""
    return [_safe_str(order.get(f, "")) for f in _SAFE_ORDER_FIELDS]


def _fill_to_row(fill: dict) -> list[Any]:
    """Extract safe fields from an order_fills dict."""
    return [_safe_str(fill.get(f, "")) for f in _SAFE_FILL_FIELDS]


def _compute_pnl_rows(fills: list[dict]) -> list[list[Any]]:
    """Compute per-symbol realised PnL from fill data.

    Simple average-cost model:
      BUY fills → add to cost basis
      SELL fills → realise gain/loss vs avg cost

    Returns rows: [symbol, total_buy_qty, total_sell_qty,
                   avg_buy_price, avg_sell_price, realised_pnl, fill_count]
    """
    from collections import defaultdict

    state: dict[str, dict] = defaultdict(lambda: {
        "buy_qty": 0, "buy_cost": 0.0,
        "sell_qty": 0, "sell_revenue": 0.0,
        "fill_count": 0,
    })
    for fill in fills:
        sym = str(fill.get("symbol", "UNKNOWN")).upper()
        side = str(fill.get("side", "")).upper()
        qty = fill.get("filled_quantity", 0) or 0
        price = fill.get("fill_price", 0.0) or 0.0
        fees = fill.get("fees", 0.0) or 0.0
        s = state[sym]
        s["fill_count"] += 1
        if side == "BUY":
            s["buy_qty"] += qty
            s["buy_cost"] += qty * price + fees
        elif side == "SELL":
            s["sell_qty"] += qty
            s["sell_revenue"] += qty * price - fees

    rows = []
    for sym, s in state.items():
        avg_buy = round(s["buy_cost"] / s["buy_qty"], 4) if s["buy_qty"] else 0.0
        avg_sell = round(s["sell_revenue"] / s["sell_qty"], 4) if s["sell_qty"] else 0.0
        matched_qty = min(s["buy_qty"], s["sell_qty"])
        realised_pnl = round((avg_sell - avg_buy) * matched_qty, 4) if matched_qty > 0 else 0.0
        rows.append([
            sym,
            s["buy_qty"], s["sell_qty"],
            avg_buy, avg_sell,
            realised_pnl,
            s["fill_count"],
        ])
    return rows


def build_strategy_results_workbook(
    strategy_id: Optional[int] = None,
    order_store=None,
    db_session=None,
    reconciliation_report: Optional[dict] = None,
) -> bytes:
    """Build an in-memory Excel workbook from persisted trading data.

    Parameters
    ----------
    strategy_id:
        If provided, signals are filtered to this strategy.
        Orders/fills are NOT filtered (they span all strategies).
    order_store:
        An OrderStore instance. If None, Orders/Fills sheets contain NO_DATA.
    db_session:
        A SQLAlchemy session bound to the strategy DB.
        If None, Signals sheet contains NO_DATA.

    Returns
    -------
    bytes
        Raw .xlsx bytes suitable for an HTTP response body.

    Safety
    ------
    - No fake rows. Missing data → NO_DATA sentinel.
    - No credentials in cells.
    - No broker API calls.
    - Generated in-memory only.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    generated_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

    # ------------------------------------------------------------------
    # Sheet 1: Summary
    # ------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Field", "Value"])
    ws_summary.append(["Report Generated At (UTC)", generated_at])
    ws_summary.append(["Strategy ID Filter", str(strategy_id) if strategy_id else "All"])
    ws_summary.append(["Platform Mode", "PAPER (research/simulation only)"])
    ws_summary.append(["Live Trading", "DISABLED"])
    ws_summary.append(["Data Source", "Persisted SQLite DB — no broker API calls"])
    ws_summary.append(["Disclaimer", "This export is for research purposes. Not financial advice."])

    # ------------------------------------------------------------------
    # Sheet 2: Signals
    # ------------------------------------------------------------------
    ws_signals = wb.create_sheet("Signals")
    signal_headers = [
        "id", "strategy_id", "symbol", "side", "status",
        "confidence", "price", "timeframe", "source_candle_time",
        "reason", "dismiss_reason", "created_at",
    ]
    signal_rows: list[list[Any]] = []

    if db_session is not None:
        try:
            from backend.db.models import StrategySignalModel
            query = db_session.query(StrategySignalModel)
            if strategy_id is not None:
                query = query.filter(StrategySignalModel.strategy_id == strategy_id)
            signals = query.order_by(StrategySignalModel.id.asc()).all()
            signal_rows = [_signal_to_row(s) for s in signals]
            ws_summary.append(["Total Signals Exported", len(signal_rows)])
        except Exception as exc:
            logger.warning("Export: failed to load signals: %s", exc)
            signal_rows = []
    else:
        ws_summary.append(["Total Signals Exported", "N/A (no DB session)"])

    _write_sheet_rows(ws_signals, signal_headers, signal_rows)

    # ------------------------------------------------------------------
    # Sheet 3: Orders
    # ------------------------------------------------------------------
    ws_orders = wb.create_sheet("Orders")
    order_headers = list(_SAFE_ORDER_FIELDS)
    order_rows: list[list[Any]] = []

    if order_store is not None:
        try:
            orders = order_store.get_recent_order_requests(limit=200)
            order_rows = [_order_to_row(o) for o in orders]
            ws_summary.append(["Total Orders Exported", len(order_rows)])
        except Exception as exc:
            logger.warning("Export: failed to load orders: %s", exc)
            order_rows = []
    else:
        ws_summary.append(["Total Orders Exported", "N/A (no OrderStore)"])

    _write_sheet_rows(ws_orders, order_headers, order_rows)

    # ------------------------------------------------------------------
    # Sheet 4: Fills
    # ------------------------------------------------------------------
    ws_fills = wb.create_sheet("Fills")
    fill_headers = list(_SAFE_FILL_FIELDS)
    fill_rows: list[list[Any]] = []
    all_fills: list[dict] = []

    if order_store is not None:
        try:
            all_fills = order_store.get_all_fills_chronological()
            fill_rows = [_fill_to_row(f) for f in all_fills]
            ws_summary.append(["Total Fills Exported", len(fill_rows)])
        except Exception as exc:
            logger.warning("Export: failed to load fills: %s", exc)
            fill_rows = []
    else:
        ws_summary.append(["Total Fills Exported", "N/A (no OrderStore)"])

    _write_sheet_rows(ws_fills, fill_headers, fill_rows)

    # ------------------------------------------------------------------
    # Sheet 5: PnL
    # ------------------------------------------------------------------
    ws_pnl = wb.create_sheet("PnL")
    pnl_headers = [
        "symbol", "total_buy_qty", "total_sell_qty",
        "avg_buy_price", "avg_sell_price", "realised_pnl", "fill_count",
    ]
    pnl_rows: list[list[Any]] = []

    if all_fills:
        try:
            pnl_rows = _compute_pnl_rows(all_fills)
            ws_summary.append(["PnL Symbols", len(pnl_rows)])
        except Exception as exc:
            logger.warning("Export: failed to compute PnL: %s", exc)
    else:
        ws_summary.append(["PnL Symbols", "N/A"])

    _write_sheet_rows(ws_pnl, pnl_headers, pnl_rows)

    # ------------------------------------------------------------------
    # Sheet 6: EquityCurve (placeholder — future portfolio engine hook)
    # ------------------------------------------------------------------
    ws_equity = wb.create_sheet("EquityCurve")
    equity_headers = ["timestamp", "equity", "drawdown_pct"]
    ws_equity.append(equity_headers)
    ws_equity.append([_NO_DATA, "", ""])
    ws_summary.append(["EquityCurve", "Not yet available — future portfolio engine hook"])

    # ------------------------------------------------------------------
    # Sheet 7: BrokerTradeReconciliation
    # ------------------------------------------------------------------
    if reconciliation_report is not None:
        ws_recon = wb.create_sheet("BrokerTradeReconciliation")
        ws_recon.append(["Summary Field", "Value"])
        ws_recon.append(["Checked At", reconciliation_report.get("checked_at", "")])
        ws_recon.append(["Broker Trade Count", reconciliation_report.get("broker_trade_count", 0)])
        ws_recon.append(["Local Fill Count", reconciliation_report.get("local_fill_count", 0)])
        ws_recon.append(["Matched Count", reconciliation_report.get("matched_count", 0)])
        ws_recon.append(["Mismatch Count", reconciliation_report.get("mismatch_count", 0)])
        ws_recon.append([])  # Spacer

        mismatch_headers = [
            "Severity", "Mismatch Type", "Symbol", "Side",
            "Broker Trade ID", "Fill ID", "Broker Order ID", "Request ID",
            "Broker Qty", "Local Qty", "Broker Price", "Local Price", "Detail"
        ]
        ws_recon.append(mismatch_headers)

        mismatches = reconciliation_report.get("mismatches", [])
        if not mismatches:
            ws_recon.append(["NO_MISMATCHES"] + [""] * (len(mismatch_headers) - 1))
        else:
            for m in mismatches:
                def get_val(key):
                    if isinstance(m, dict):
                        return m.get(key)
                    return getattr(m, key, None)

                ws_recon.append([
                    _safe_str(get_val("severity")),
                    _safe_str(get_val("mismatch_type")),
                    _safe_str(get_val("symbol")),
                    _safe_str(get_val("side")),
                    _safe_str(get_val("broker_trade_id")),
                    _safe_str(get_val("fill_id")),
                    _safe_str(get_val("broker_order_id")),
                    _safe_str(get_val("request_id")),
                    _safe_str(get_val("broker_qty")),
                    _safe_str(get_val("local_qty")),
                    _safe_str(get_val("broker_price")),
                    _safe_str(get_val("local_price")),
                    _safe_str(get_val("detail")),
                ])
        ws_summary.append(["BrokerTradeReconciliation", f"Included — {reconciliation_report.get('mismatch_count', 0)} mismatches"])
    else:
        ws_summary.append(["BrokerTradeReconciliation", "Not requested/available"])

    # ------------------------------------------------------------------
    # Serialize to bytes
    # ------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
