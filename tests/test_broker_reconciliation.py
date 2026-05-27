# tests/test_broker_reconciliation.py
"""
Tests for Broker Trade Book Reconciliation to Fill Ledger — Phase 22B

ABSOLUTE SAFETY: All tests use mocks. Zero real broker API calls.
No live orders placed. No credentials printed.
"""

import pytest
from datetime import datetime, timezone
from unittest.mock import MagicMock, patch
from fastapi.testclient import TestClient
from fastapi import FastAPI

from backend.services.broker_trade_reconciliation import (
    BrokerTradeReconciliationService,
    BrokerTradeSnapshot,
    InternalFillSnapshot,
    TradeReconciliationMismatch,
    TradeReconciliationReport,
)
from backend.routers.broker_account import router as broker_account_router
from backend.services.strategy_export_service import build_strategy_results_workbook
from openpyxl import load_workbook
import io

class MockOrderStore:
    def __init__(self, fills):
        self._fills = fills

    def get_all_fills_chronological(self):
        return self._fills


def test_reconcile_empty():
    """Verify empty input behaves correctly."""
    service = BrokerTradeReconciliationService()
    report = service.reconcile_trades(broker_trades=[], internal_fills=[])
    
    assert report.broker_trade_count == 0
    assert report.local_fill_count == 0
    assert report.matched_count == 0
    assert report.mismatch_count == 0
    assert len(report.mismatches) == 0


def test_reconcile_exact_match_by_trade_id():
    """Verify exact match by trade ID is reconciled."""
    fill = {
        "fill_id": "FILL123",
        "request_id": "REQ123",
        "broker_order_id": "BORD123",
        "symbol": "TCS",
        "side": "BUY",
        "filled_quantity": 10,
        "fill_price": 3200.0,
        "created_at": "2026-05-27T10:00:00Z"
    }

    broker_trades = [
        {
            "tradeid": "FILL123",
            "orderid": "BORD123",
            "tradingsymbol": "TCS",
            "transactiontype": "BUY",
            "quantity": 10,
            "tradeprice": 3200.0,
            "updatetime": "2026-05-27 10:00:00"
        }
    ]

    service = BrokerTradeReconciliationService()
    report = service.reconcile_trades(broker_trades, [fill])
    assert report.broker_trade_count == 1
    assert report.local_fill_count == 1
    assert report.matched_count == 1
    assert report.mismatch_count == 0
    assert len(report.mismatches) == 0


def test_reconcile_quantity_mismatch():
    """Verify quantity mismatch is reported as MEDIUM."""
    fill = {
        "fill_id": "FILL123",
        "request_id": "REQ123",
        "broker_order_id": "BORD123",
        "symbol": "TCS",
        "side": "BUY",
        "filled_quantity": 10,
        "fill_price": 3200.0,
        "created_at": "2026-05-27T10:00:00Z"
    }

    broker_trades = [
        {
            "tradeid": "FILL123",
            "orderid": "BORD123",
            "tradingsymbol": "TCS",
            "transactiontype": "BUY",
            "quantity": 8,  # Mismatch
            "tradeprice": 3200.0,
            "updatetime": "2026-05-27 10:00:00"
        }
    ]

    service = BrokerTradeReconciliationService()
    report = service.reconcile_trades(broker_trades, [fill])
    assert report.broker_trade_count == 1
    assert report.matched_count == 1  # Note: it's matched on ID, but has mismatch records
    assert report.mismatch_count == 1
    mismatch = report.mismatches[0]
    assert mismatch.severity == "MEDIUM"
    assert mismatch.mismatch_type == "QUANTITY_MISMATCH"
    assert mismatch.broker_qty == 8
    assert mismatch.local_qty == 10


def test_reconcile_price_mismatch():
    """Verify price mismatch is reported as MEDIUM."""
    fill = {
        "fill_id": "FILL123",
        "request_id": "REQ123",
        "broker_order_id": "BORD123",
        "symbol": "TCS",
        "side": "BUY",
        "filled_quantity": 10,
        "fill_price": 3200.0,
        "created_at": "2026-05-27T10:00:00Z"
    }

    broker_trades = [
        {
            "tradeid": "FILL123",
            "orderid": "BORD123",
            "tradingsymbol": "TCS",
            "transactiontype": "BUY",
            "quantity": 10,
            "tradeprice": 3250.0,  # Price mismatch
            "updatetime": "2026-05-27 10:00:00"
        }
    ]

    service = BrokerTradeReconciliationService()
    report = service.reconcile_trades(broker_trades, [fill])
    assert report.broker_trade_count == 1
    assert report.matched_count == 1
    assert report.mismatch_count == 1
    mismatch = report.mismatches[0]
    assert mismatch.severity == "MEDIUM"
    assert mismatch.mismatch_type == "PRICE_MISMATCH"
    assert mismatch.broker_price == 3250.0
    assert mismatch.local_price == 3200.0


def test_reconcile_ghost_broker_trade():
    """Verify broker trade that doesn't match any internal fill is HIGH."""
    broker_trades = [
        {
            "tradeid": "BFILL_GHOST",
            "orderid": "BORD999",
            "tradingsymbol": "INFY",
            "transactiontype": "SELL",
            "quantity": 100,
            "tradeprice": 1400.0,
            "updatetime": "2026-05-27 10:00:00"
        }
    ]

    service = BrokerTradeReconciliationService()
    report = service.reconcile_trades(broker_trades, [])
    assert report.broker_trade_count == 1
    assert report.matched_count == 0
    assert report.mismatch_count == 1
    mismatch = report.mismatches[0]
    assert mismatch.severity == "HIGH"
    assert mismatch.mismatch_type == "BROKER_TRADE_MISSING_LOCAL_FILL"
    assert mismatch.broker_trade_id == "BFILL_GHOST"


def test_reconcile_matching_proximity_no_trade_id():
    """Verify proximity matching works when trade/fill IDs are not linked."""
    fill = {
        "fill_id": "FILL_PROX",
        "request_id": "REQ_PROX",
        "broker_order_id": None,
        "symbol": "SBIN",
        "side": "BUY",
        "filled_quantity": 50,
        "fill_price": 600.0,
        "created_at": "2026-05-27T10:00:00Z"
    }

    # Broker trade close in time, price, qty
    broker_trades = [
        {
            "tradeid": "BFILL_PROX",
            "orderid": "BORD_PROX",
            "tradingsymbol": "SBIN",
            "transactiontype": "BUY",
            "quantity": 50,
            "tradeprice": 600.01,  # within price and time tolerance
            "updatetime": "2026-05-27 10:00:10"
        }
    ]

    service = BrokerTradeReconciliationService()
    report = service.reconcile_trades(broker_trades, [fill])
    assert report.broker_trade_count == 1
    assert report.matched_count == 1
    assert report.mismatch_count == 0


def test_reconcile_side_mismatch():
    """Verify transaction side mismatch prevents match and causes unmatched."""
    fill = {
        "fill_id": "FILL123",
        "request_id": "REQ123",
        "broker_order_id": "BORD123",
        "symbol": "TCS",
        "side": "BUY",
        "filled_quantity": 10,
        "fill_price": 3200.0,
        "created_at": "2026-05-27T10:00:00Z"
    }

    # Broker trade with different side
    broker_trades = [
        {
            "tradeid": "FILL123",
            "orderid": "BORD123",
            "tradingsymbol": "TCS",
            "transactiontype": "SELL",
            "quantity": 10,
            "tradeprice": 3200.0,
            "updatetime": "2026-05-27 10:00:00"
        }
    ]

    service = BrokerTradeReconciliationService()
    report = service.reconcile_trades(broker_trades, [fill])
    assert report.matched_count == 0
    # Since they have same broker_order_id, it matches in PASS 1 but reports SIDE_MISMATCH
    assert any(m.mismatch_type == "SIDE_MISMATCH" for m in report.mismatches)


# =====================================================================
# API / Router Tests
# =====================================================================

def test_api_reconcile_and_report_endpoints():
    """Verify POST /reconcile triggers reconciliation and GET /reconciliation-report returns it."""
    app = FastAPI()
    app.include_router(broker_account_router, prefix="/api")
    
    mock_order_store = MockOrderStore([])
    app.state.order_store = mock_order_store
    
    from backend.core.security import require_admin_token
    app.dependency_overrides[require_admin_token] = lambda: {"admin": True}

    client = TestClient(app)

    # Set up mock session manager to look valid
    mock_sm = MagicMock()
    mock_sm.is_valid = True
    app.state.session_manager = mock_sm

    mock_report = TradeReconciliationReport(
        checked_at="2026-05-27T10:00:00Z",
        broker_trade_count=1,
        local_fill_count=0,
        matched_count=0,
        mismatch_count=1,
        mismatches=[
            TradeReconciliationMismatch(
                severity="HIGH",
                mismatch_type="BROKER_TRADE_MISSING_LOCAL_FILL",
                symbol="RELIANCE",
                side="BUY",
                broker_trade_id="BFILL_TEST",
                broker_qty=10,
                broker_price=2500.0,
            )
        ]
    )

    with patch("backend.services.broker_trade_reconciliation.BrokerTradeReconciliationService.reconcile_from_broker", return_value=mock_report):
        # 1. Reconcile
        response = client.post("/api/broker/account/trade-reconciliation/run")
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "OK"
        report_data = data["report"]
        assert report_data["checked_at"] == "2026-05-27T10:00:00Z"
        assert report_data["broker_trade_count"] == 1
        assert report_data["mismatch_count"] == 1

        # 2. Get Report
        get_response = client.get("/api/broker/account/trade-reconciliation/status")
        assert get_response.status_code == 200
        get_data = get_response.json()
        assert get_data["status"] == "OK"
        assert get_data["last_run"]["broker_trade_count"] == 1


# =====================================================================
# Strategy Export Integration
# =====================================================================

def test_strategy_export_reconciliation_sheet():
    """Verify that exporting workbook generates BrokerTradeReconciliation sheet when available."""
    recon_report = {
        "checked_at": "2026-05-27T12:00:00Z",
        "broker_trade_count": 2,
        "local_fill_count": 1,
        "matched_count": 1,
        "mismatch_count": 1,
        "mismatches": [
            {
                "severity": "MEDIUM",
                "mismatch_type": "PRICE_MISMATCH",
                "symbol": "SBIN",
                "side": "BUY",
                "broker_trade_id": "BFILL_1",
                "fill_id": "LFILL_1",
                "broker_qty": 10,
                "local_qty": 10,
                "broker_price": 605.0,
                "local_price": 600.0,
                "detail": "Price diff: 5.0",
            }
        ]
    }

    # Generate workbook bytes
    xlsx_bytes = build_strategy_results_workbook(
        strategy_id=None,
        order_store=MockOrderStore([]),
        db_session=None,
        reconciliation_report=recon_report,
    )

    # Load with openpyxl
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    assert "BrokerTradeReconciliation" in wb.sheetnames
    
    ws = wb["BrokerTradeReconciliation"]
    
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][:2] == ("Summary Field", "Value")
    assert rows[1][:2] == ("Checked At", "2026-05-27T12:00:00Z")
    assert rows[2][:2] == ("Broker Trade Count", 2)
    assert rows[3][:2] == ("Local Fill Count", 1)
    assert rows[4][:2] == ("Matched Count", 1)
    assert rows[5][:2] == ("Mismatch Count", 1)
    
    # Verify mismatch row
    assert rows[7][:13] == (
        "Severity", "Mismatch Type", "Symbol", "Side",
        "Broker Trade ID", "Fill ID", "Broker Order ID", "Request ID",
        "Broker Qty", "Local Qty", "Broker Price", "Local Price", "Detail"
    )
    
    mismatch_row = rows[8]
    assert mismatch_row[0] == "MEDIUM"
    assert mismatch_row[1] == "PRICE_MISMATCH"
    assert mismatch_row[2] == "SBIN"
    assert mismatch_row[4] == "BFILL_1"
    assert mismatch_row[5] == "LFILL_1"
    assert int(mismatch_row[8]) == 10
    assert float(mismatch_row[10]) == 605.0
    assert float(mismatch_row[11]) == 600.0
